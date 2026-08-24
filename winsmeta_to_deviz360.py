# -*- coding: utf-8 -*-
"""Export Winsmeta (.KOS) -> Deviz360 (.xlsx)."""

from __future__ import annotations

import os
import re
import unicodedata
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from winsmeta_to_devizonline import (
    DEVIZ_COLUMNS,
    KosData,
    ProjectInfo,
    _collect_q_items,
    _empty_row,
    _export_norm_block,
    _pick_name,
    _top_level_chapters,
    load_kos,
)

DEFAULT_TVA_PCT = 20

CENTRALIZATOR_COLUMNS = [
    "Foaie",
    "Simbol",
    "Denumire",
    "Cantitate",
    "UM",
    "TVA",
    "Data",
    "PozitieDG",
    "PozitieGantt",
    "DataInceput",
    "DataSfarsit",
    "DataInceputEch",
    "DataSfarsitEch",
]

D360_ZERO = "0.00000000000000000"
D360_CURS_EURO = "20.0536"


def _d360_num(value: Any) -> str:
    if value is None or value == "":
        return D360_ZERO
    try:
        return f"{float(value):.17f}"
    except (TypeError, ValueError):
        return str(value)


def _d360_int_str(value: Any) -> str:
    if value is None or value == "":
        return "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    cleaned = cleaned.replace("\r", "").replace("\n", "").strip()
    if not cleaned:
        cleaned = "Deviz"
    return cleaned[:31]


def _sheet_code(index: int, title: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9_]+", "_", title)[:24].strip("_")
    return _sanitize_sheet_name(f"D{index:02d}_{slug or 'Deviz'}")


DEVIZ_TYPE_ORDER = {"constructie": 1, "montare": 2, "utilaj": 3}


def _normalize_chapter_name(name: str) -> str:
    text = unicodedata.normalize("NFKD", str(name or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def _deviz_type_name(name: str) -> Optional[str]:
    """Recunoaște capitole tip deviz Deviz360 (construcție / montare / utilaj)."""
    n = _normalize_chapter_name(name)
    if not n:
        return None
    if re.search(r"lucrari\s+de\s+constructii?", n):
        return "constructie"
    if re.search(r"lucrari\s+de\s+montare", n) or n == "montare":
        return "montare"
    if n == "utilaj" or n.startswith("utilaj "):
        return "utilaj"
    return None


def _partition_deviz_chapters(data: KosData) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Împarte capitolele de nivel 1: tip deviz standard vs rest (încăperi etc.)."""
    typed: List[Dict[str, Any]] = []
    other: List[Dict[str, Any]] = []
    for chapter in _top_level_chapters(data.pozycje):
        if not _collect_q_items(data.pozycje, chapter["Nr"]):
            continue
        name = _pick_name(chapter.get("Nazwa"))
        if _deviz_type_name(name):
            typed.append(chapter)
        else:
            other.append(chapter)
    typed.sort(
        key=lambda c: DEVIZ_TYPE_ORDER.get(
            _deviz_type_name(_pick_name(c.get("Nazwa"))) or "", 99
        )
    )
    return typed, other


def _deviz_chapters(data: KosData) -> List[Dict[str, Any]]:
    """Capitole tip deviz standard (construcție / montare / utilaj) cu norme."""
    typed, _ = _partition_deviz_chapters(data)
    return typed


def _format_section_denumire(number: str, name: str) -> str:
    """Format Winsmeta: «1.1. Tablou general...»."""
    number = str(number or "").strip().rstrip(".")
    name = str(name or "").strip()
    if number and name:
        return f"{number}. {name}"
    if number:
        return f"{number}."
    return name


def _combine_section_names(*parts: str) -> str:
    """Unește denumirile ierarhice: «Demolari - Cota 0.000»."""
    seen: set[str] = set()
    cleaned: List[str] = []
    for part in parts:
        text = str(part or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return " - ".join(cleaned)


def _section_display_name(
    name: str,
    parent_names: Tuple[str, ...] = (),
    merge_parent_names: bool = False,
) -> str:
    if merge_parent_names and parent_names:
        return _combine_section_names(*parent_names, name)
    return name


def _root_chapter_number(data: KosData, chapter_nr: int) -> str:
    """Indexul capitolului de nivel 1 în arborele POZYCJE (1, 2, 3...)."""
    root = data.pozycje.get(0)
    if not root:
        return ""
    idx = 0
    child = root.get("NrPod") or 0
    while child:
        poz = data.pozycje.get(child)
        if poz and poz.get("Typ") == "S":
            idx += 1
            if poz["Nr"] == chapter_nr:
                return str(idx)
        child = poz.get("NrNast") if poz else 0
    return ""


def _deviz_header_title(info: ProjectInfo, chapter_name: str | None = None) -> str:
    if chapter_name:
        name = chapter_name.strip()
        if name.lower().startswith("deviz"):
            return name
        return f"Deviz {name}"
    deviz = str(info.deviz or "").strip()
    if not deviz:
        return "Deviz"
    if deviz.lower().startswith("deviz"):
        return deviz
    return f"Deviz {deviz}"


def _centralizator_denumire(deviz_title: str) -> str:
    title = deviz_title.strip()
    if title.lower().startswith("deviz "):
        return title[6:].strip() or title
    return title


def _fill(out: Dict[str, Any], key: str, value: Any) -> None:
    if out.get(key) in ("", None):
        out[key] = value


def _format_money(value: Any) -> str:
    if value in ("", None):
        return "0"
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if num == 0:
        return "0"
    text = f"{num:.10f}".rstrip("0").rstrip(".")
    return text or "0"


def _header_row_d360(text: str) -> Dict[str, Any]:
    row = {col: None for col in DEVIZ_COLUMNS}
    row["Simbol"] = "#R"
    row["Denumire"] = text
    return row


def _format_row_d360(row: Dict[str, Any], info: ProjectInfo) -> Dict[str, Any]:
    out = _empty_row()
    out.update(row)

    tip = out.get("Tip")
    if tip is not None and tip != "":
        if isinstance(tip, float) and tip.is_integer():
            out["Tip"] = str(int(tip))
        else:
            out["Tip"] = str(tip)

    tip_str = out.get("Tip")
    if tip_str not in ("0", "1", "2", "3", "100", "101"):
        return out

    if out.get("Pozitie") not in ("", None):
        out["Pozitie"] = _d360_int_str(out["Pozitie"])

    out["PozitieGrup"] = _d360_int_str(out.get("PozitieGrup", 0))

    for col in ("CantitateArticol", "CantitateinArticol", "CantitateTotala"):
        if out.get(col) not in ("", None):
            out[col] = _d360_num(out[col])

    for col in ("PretMat", "PretMan", "PretUti", "PretTr", "PretTotal"):
        if out.get(col) not in ("", None):
            out[col] = _format_money(out[col])

    if tip_str in ("1", "2", "3") and out.get("Pret") not in ("", None):
        out["Pret"] = _d360_num(out["Pret"])

    _fill(out, "MonedaPretAuto", info.moneda)
    _fill(out, "Moneda", info.moneda)
    _fill(out, "TipPret", "1")
    _fill(out, "CantitateDeviz", "1")
    _fill(out, "TipUtilaj", "a")
    _fill(out, "SporMan", "0")
    _fill(out, "SporMat", "0")
    _fill(out, "SporUti", "0")
    _fill(out, "Distanta", D360_ZERO)
    _fill(out, "CursEuroDeviz", D360_CURS_EURO)
    _fill(out, "Greutate", D360_ZERO)
    _fill(out, "codRecapitulatie", "0")
    _fill(out, "codProvenienta", "0")
    _fill(out, "CantitateBeneficiar", "0")

    ca = out.get("codArticol")
    if ca in ("", None, 0, 0.0):
        out["codArticol"] = "0"
    else:
        out["codArticol"] = _d360_int_str(ca)

    if tip_str == "0":
        _fill(out, "Pret", "0")
        _fill(out, "PretAuto", "0")
        _fill(out, "Provenienta", "A")
    elif tip_str in ("1", "2", "3"):
        _fill(out, "Provenienta", "AR")
        _fill(out, "PretAuto", D360_ZERO)
    elif tip_str in ("100", "101"):
        _fill(out, "Provenienta", "C")
        _fill(out, "CantitateArticol", D360_ZERO)
        _fill(out, "CantitateinArticol", D360_ZERO)
        _fill(out, "Pret", D360_ZERO)
        _fill(out, "PretAuto", D360_ZERO)
        if tip_str == "101":
            _fill(out, "CantitateTotala", D360_ZERO)

    _fill(out, "PretMat", "0")
    _fill(out, "PretMan", "0")
    _fill(out, "PretUti", "0")
    _fill(out, "PretTr", "0")
    if tip_str != "101" or out.get("PretTotal") in ("", None):
        _fill(out, "PretTotal", "0")

    out["TipPret"] = _d360_int_str(out.get("TipPret", 1))
    out["CantitateDeviz"] = "1"
    out["PozitieGrup"] = _d360_int_str(out.get("PozitieGrup", 0))

    return out


def _sheet_write_value(value: Any) -> Any:
    if value == "" or value is None:
        return None
    if isinstance(value, bool) or isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text.startswith("#"):
        return value
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except ValueError:
            return value
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return float(text)
        except ValueError:
            return value
    return value


def _section_row_d360(
    name: str,
    pozitie: int,
    qty_total: float,
    info: ProjectInfo,
    section_number: str = "",
) -> Dict[str, Any]:
    row = _empty_row()
    row["Tip"] = "100"
    row["Pozitie"] = pozitie
    row["PozitieGrup"] = "0"
    row["Denumire"] = _format_section_denumire(section_number, name)
    row["CantitateTotala"] = qty_total
    row["MonedaPretAuto"] = info.moneda
    row["Moneda"] = info.moneda
    row["TipPret"] = "1"
    row["Provenienta"] = "C"
    return _format_row_d360(row, info)


def _chapter_total_row_d360(
    name: str,
    pozitie: int,
    mat: float,
    man: float,
    uti: float,
    total: float,
    info: ProjectInfo,
    section_number: str = "",
) -> Dict[str, Any]:
    title = _format_section_denumire(section_number, name)
    row = _empty_row()
    row["Tip"] = "101"
    row["Pozitie"] = pozitie
    row["PozitieGrup"] = "0"
    row["Denumire"] = f"TOTAL {title}"
    row["PretMat"] = mat
    row["PretMan"] = man
    row["PretUti"] = uti
    row["PretTotal"] = total
    row["MonedaPretAuto"] = info.moneda
    row["Moneda"] = info.moneda
    row["TipPret"] = "1"
    row["Provenienta"] = "C"
    return _format_row_d360(row, info)


def _format_block_d360(block: List[Dict[str, Any]], info: ProjectInfo) -> List[Dict[str, Any]]:
    return [_format_row_d360(r, info) for r in block]


def _emit_section_branch(
    data: KosData,
    nr: int,
    rows: List[Dict[str, Any]],
    poz_counter: List[int],
    totals: Dict[str, float],
    section_number: str = "",
    parent_names: Tuple[str, ...] = (),
    merge_parent_names: bool = False,
) -> None:
    poz = data.pozycje.get(nr)
    if not poz or poz.get("Typ") != "S":
        return

    section_rows: List[Dict[str, Any]] = []
    sec_mat = sec_man = sec_uti = sec_mh = sec_total = sec_qty = 0.0
    first_poz = last_poz = 0
    nested_idx = 0

    child = poz.get("NrPod") or 0
    while child:
        child_poz = data.pozycje.get(child)
        if not child_poz:
            break
        if child_poz.get("Typ") == "S":
            nested_idx += 1
            nested_number = (
                f"{section_number}.{nested_idx}" if section_number else str(nested_idx)
            )
            nested_parents = parent_names
            if merge_parent_names:
                nested_parents = parent_names + (_pick_name(poz.get("Nazwa")),)
            _emit_section_branch(
                data,
                child,
                rows,
                poz_counter,
                totals,
                nested_number,
                parent_names=nested_parents,
                merge_parent_names=merge_parent_names,
            )
        elif child_poz.get("Typ") == "Q":
            poz_counter[0] += 1
            pozitie = poz_counter[0]
            if not first_poz:
                first_poz = pozitie
            last_poz = pozitie
            block, mat, man, uti, mh, line_total = _export_norm_block(data, child_poz, pozitie)
            section_rows.extend(block)
            sec_mat += mat
            sec_man += man
            sec_uti += uti
            sec_mh += mh
            sec_total += line_total
            sec_qty += float(child_poz.get("Ilosc") or 0.0)
            totals["mat"] += mat
            totals["man"] += man
            totals["uti"] += uti
            totals["mh"] += mh
            totals["total"] += line_total
        child = child_poz.get("NrNast") or 0

    if not section_rows:
        return

    name = _pick_name(poz.get("Nazwa"))
    display_name = _section_display_name(name, parent_names, merge_parent_names)
    rows.append(
        _section_row_d360(display_name, first_poz, sec_qty, data.info, section_number)
    )
    rows.extend(_format_block_d360(section_rows, data.info))
    rows.append(
        _chapter_total_row_d360(
            display_name,
            last_poz,
            sec_mat,
            sec_man,
            sec_uti,
            sec_total,
            data.info,
            section_number,
        )
    )


def _emit_chapter_as_section(
    data: KosData,
    chapter_name: str,
    q_items: List[Dict[str, Any]],
    rows: List[Dict[str, Any]],
    poz_counter: List[int],
    totals: Dict[str, float],
    section_number: str = "",
) -> None:
    section_rows: List[Dict[str, Any]] = []
    sec_mat = sec_man = sec_uti = sec_mh = sec_total = sec_qty = 0.0
    first_poz = last_poz = 0

    for poz in q_items:
        poz_counter[0] += 1
        pozitie = poz_counter[0]
        if not first_poz:
            first_poz = pozitie
        last_poz = pozitie
        block, mat, man, uti, mh, line_total = _export_norm_block(data, poz, pozitie)
        section_rows.extend(block)
        sec_mat += mat
        sec_man += man
        sec_uti += uti
        sec_mh += mh
        sec_total += line_total
        sec_qty += float(poz.get("Ilosc") or 0.0)
        totals["mat"] += mat
        totals["man"] += man
        totals["uti"] += uti
        totals["mh"] += mh
        totals["total"] += line_total

    if not section_rows:
        return

    rows.append(
        _section_row_d360(chapter_name, first_poz, sec_qty, data.info, section_number)
    )
    rows.extend(_format_block_d360(section_rows, data.info))
    rows.append(
        _chapter_total_row_d360(
            chapter_name,
            last_poz,
            sec_mat,
            sec_man,
            sec_uti,
            sec_total,
            data.info,
            section_number,
        )
    )


def _emit_chapter_content(
    data: KosData,
    chapter_nr: int,
    rows: List[Dict[str, Any]],
    poz_counter: List[int],
    totals: Dict[str, float],
    chapter_number: str = "",
    merge_parent_names: bool = False,
) -> None:
    chapter = data.pozycje.get(chapter_nr)
    if not chapter:
        return

    chapter_name = _pick_name(chapter.get("Nazwa"))
    parent_names = (chapter_name,) if merge_parent_names and chapter_name else ()

    direct_q: List[Dict[str, Any]] = []
    section_idx = 0
    child = chapter.get("NrPod") or 0
    while child:
        child_poz = data.pozycje.get(child)
        if not child_poz:
            break
        if child_poz.get("Typ") == "S":
            section_idx += 1
            section_number = (
                f"{chapter_number}.{section_idx}" if chapter_number else str(section_idx)
            )
            _emit_section_branch(
                data,
                child,
                rows,
                poz_counter,
                totals,
                section_number,
                parent_names=parent_names,
                merge_parent_names=merge_parent_names,
            )
        elif child_poz.get("Typ") == "Q":
            direct_q.append(child_poz)
        child = child_poz.get("NrNast") or 0

    if direct_q:
        _emit_chapter_as_section(
            data,
            chapter_name,
            direct_q,
            rows,
            poz_counter,
            totals,
            chapter_number,
        )


def _emit_root_content(
    data: KosData,
    rows: List[Dict[str, Any]],
    poz_counter: List[int],
    totals: Dict[str, float],
) -> None:
    """Parcurge tot proiectul de la rădăcină (un singur deviz)."""
    root = data.pozycje.get(0)
    if not root:
        return

    direct_q: List[Dict[str, Any]] = []
    child = root.get("NrPod") or 0
    while child:
        child_poz = data.pozycje.get(child)
        if not child_poz:
            break
        if child_poz.get("Typ") == "S":
            chapter_number = _root_chapter_number(data, child)
            _emit_chapter_content(
                data, child, rows, poz_counter, totals, chapter_number, True
            )
        elif child_poz.get("Typ") == "Q":
            direct_q.append(child_poz)
        child = child_poz.get("NrNast") or 0

    if direct_q:
        _emit_chapter_as_section(data, "Deviz", direct_q, rows, poz_counter, totals)


def _emit_chapters_content(
    data: KosData,
    chapters: List[Dict[str, Any]],
    rows: List[Dict[str, Any]],
    poz_counter: List[int],
    totals: Dict[str, float],
) -> None:
    for chapter in chapters:
        chapter_number = _root_chapter_number(data, chapter["Nr"])
        _emit_chapter_content(
            data, chapter["Nr"], rows, poz_counter, totals, chapter_number, True
        )


def convert_kos_to_chapter_sheet_rows_d360(
    data: KosData,
    chapter: Dict[str, Any],
    deviz_title: str,
) -> Tuple[List[Dict[str, Any]], int, Dict[str, float]]:
    rows: List[Dict[str, Any]] = []

    if data.info.obiectiv:
        rows.append(_header_row_d360(f"Obiectiv {data.info.obiectiv}"))
    if data.info.obiect:
        rows.append(_header_row_d360(f"Obiect {data.info.obiect}"))
    rows.append(_header_row_d360(deviz_title))

    poz_counter = [0]
    totals = {"mat": 0.0, "man": 0.0, "uti": 0.0, "mh": 0.0, "total": 0.0}
    chapter_number = _root_chapter_number(data, chapter["Nr"])
    _emit_chapter_content(data, chapter["Nr"], rows, poz_counter, totals, chapter_number)

    if poz_counter[0] == 0:
        raise ValueError(f"Capitolul «{_pick_name(chapter.get('Nazwa'))}» nu conține norme.")

    return rows, poz_counter[0], totals


def convert_kos_to_root_sheet_rows_d360(
    data: KosData,
    deviz_title: str,
    chapters: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[List[Dict[str, Any]], int, Dict[str, float]]:
    """Un deviz: tot proiectul sau capitole non-standard (încăperi) grupate."""
    rows: List[Dict[str, Any]] = []

    if data.info.obiectiv:
        rows.append(_header_row_d360(f"Obiectiv {data.info.obiectiv}"))
    if data.info.obiect:
        rows.append(_header_row_d360(f"Obiect {data.info.obiect}"))
    rows.append(_header_row_d360(deviz_title))

    poz_counter = [0]
    totals = {"mat": 0.0, "man": 0.0, "uti": 0.0, "mh": 0.0, "total": 0.0}

    if chapters:
        _emit_chapters_content(data, chapters, rows, poz_counter, totals)
    else:
        _emit_root_content(data, rows, poz_counter, totals)

    if poz_counter[0] == 0:
        q_items = _collect_q_items(data.pozycje, 0)
        for poz in q_items:
            poz_counter[0] += 1
            block, mat, man, uti, mh, line_total = _export_norm_block(
                data, poz, poz_counter[0]
            )
            rows.extend(_format_block_d360(block, data.info))
            totals["mat"] += mat
            totals["man"] += man
            totals["uti"] += uti
            totals["mh"] += mh
            totals["total"] += line_total

    if poz_counter[0] == 0:
        raise ValueError("Proiectul Winsmeta nu conține norme (POZYCJE Typ=Q).")

    return rows, poz_counter[0], totals


def _write_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    ws.append(DEVIZ_COLUMNS)
    for src in rows:
        ws.append([_sheet_write_value(src.get(col)) for col in DEVIZ_COLUMNS])


def _deviz_date(info: ProjectInfo) -> str:
    if info.data_deviz:
        return info.data_deviz[:10]
    return date.today().isoformat()


def write_deviz360_xlsx(
    sheets: List[Tuple[str, str, List[Dict[str, Any]]]],
    centralizator_rows: List[Dict[str, Any]],
    output_path: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, _, rows in sheets:
        ws = wb.create_sheet(title=_sanitize_sheet_name(sheet_name))
        _write_sheet(ws, rows)

    ws_c = wb.create_sheet(title="Centralizator")
    ws_c.append(CENTRALIZATOR_COLUMNS)
    for row in centralizator_rows:
        ws_c.append([row.get(col) for col in CENTRALIZATOR_COLUMNS])

    wb.save(output_path)


def convert_kos_to_deviz360_xlsx(
    kos_path: str,
    output_path: str,
) -> Tuple[int, ProjectInfo, int, int, float, float, float, float]:
    data = load_kos(kos_path)
    typed, other = _partition_deviz_chapters(data)
    if not typed and not other:
        raise ValueError("Proiectul Winsmeta nu conține norme (POZYCJE Typ=Q).")

    sheets: List[Tuple[str, str, List[Dict[str, Any]]]] = []
    centralizator: List[Dict[str, Any]] = []
    deviz_date = _deviz_date(data.info)
    total_rows = 0
    total_norms = 0
    sum_mat = sum_man = sum_uti = sum_total = 0.0
    sheet_idx = 0

    def _add_sheet(sheet_name: str, deviz_title: str, rows: List, norm_count: int, totals: Dict):
        nonlocal total_rows, total_norms, sum_mat, sum_man, sum_uti, sum_total, sheet_idx
        total_rows += len(rows)
        total_norms += norm_count
        sum_mat += totals["mat"]
        sum_man += totals["man"]
        sum_uti += totals["uti"]
        sum_total += totals["total"]
        sheets.append((sheet_name, deviz_title, rows))
        sheet_idx += 1
        centralizator.append(
            {
                "Foaie": sheet_name,
                "Simbol": None,
                "Denumire": _centralizator_denumire(deviz_title),
                "Cantitate": _d360_num(1),
                "UM": "buc",
                "TVA": DEFAULT_TVA_PCT,
                "Data": deviz_date,
                "PozitieDG": "0",
                "PozitieGantt": sheet_idx,
                "DataInceput": deviz_date,
                "DataSfarsit": deviz_date,
                "DataInceputEch": deviz_date,
                "DataSfarsitEch": deviz_date,
            }
        )

    if not typed:
        deviz_title = _deviz_header_title(data.info)
        sheet_name = _sheet_code(1, deviz_title)
        rows, norm_count, totals = convert_kos_to_root_sheet_rows_d360(data, deviz_title)
        _add_sheet(sheet_name, deviz_title, rows, norm_count, totals)
    else:
        for chapter in typed:
            chapter_name = _pick_name(chapter.get("Nazwa"))
            deviz_title = _deviz_header_title(data.info, chapter_name)
            sheet_name = _sheet_code(sheet_idx + 1, chapter_name)
            rows, norm_count, totals = convert_kos_to_chapter_sheet_rows_d360(
                data, chapter, deviz_title
            )
            _add_sheet(sheet_name, deviz_title, rows, norm_count, totals)

        if other:
            deviz_title = _deviz_header_title(data.info)
            sheet_name = _sheet_code(sheet_idx + 1, deviz_title)
            rows, norm_count, totals = convert_kos_to_root_sheet_rows_d360(
                data, deviz_title, chapters=other
            )
            _add_sheet(sheet_name, deviz_title, rows, norm_count, totals)

    write_deviz360_xlsx(sheets, centralizator, output_path)

    return (
        total_rows,
        data.info,
        total_norms,
        len(sheets),
        sum_mat,
        sum_man,
        sum_uti,
        sum_total,
    )


def default_output_path_deviz360(kos_path: str) -> str:
    parent = os.path.dirname(os.path.abspath(kos_path))
    return os.path.join(parent, "deviz360_export.xlsx")
